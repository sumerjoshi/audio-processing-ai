import argparse
import csv
from datetime import datetime
from pathlib import Path
from typing import Tuple, Literal
import wave

import numpy as np
import torch
import torch.nn.functional as F
import torchaudio
from torch import Tensor
from torchaudio.transforms import MelSpectrogram
from tqdm.auto import tqdm
from model.pretrained.dual_head_cnn14 import DualHeadCnn14Simple
import pandas as pd 
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from typing import Dict

try:
    import librosa
    LIBROSA_AVAILABLE = True
except ImportError:
    LIBROSA_AVAILABLE = False

timestamp = datetime.now().strftime("%Y%m%d_%H%M")
DEFAULT_CSV_PATH = f"predictions_{timestamp}.csv"


def preprocess_audio(file_path, sample_rate=16000, duration=10.0) -> Tensor:
    waveform, sr = torchaudio.load(file_path)

    if sr != sample_rate:
        waveform = torchaudio.transforms.Resample(sr, sample_rate)(waveform)

    if waveform.shape[0] > 1:
        waveform = waveform.mean(dim=0, keepdim=True)

    waveform_len = waveform.shape[1]
    target_len = int(sample_rate * duration)

    if waveform_len < target_len:
        pad = target_len - waveform_len
        waveform = F.pad(waveform, (0, pad))
    else:
        start = (waveform_len - target_len) // 2
        waveform = waveform[:, start : start + target_len]

    return waveform

def predict_one(model, input_tensor: Tensor) -> Tuple[torch.Tensor, torch.Tensor]:
    with torch.no_grad():
        if input_tensor.ndim == 2:
            input_tensor = input_tensor.squeeze(0)
        input_tensor = input_tensor.unsqueeze(0)
        binary_logit, tag_logits = model(input_tensor)
        ai_prob = torch.sigmoid(binary_logit).item()
        tag_probs = torch.sigmoid(tag_logits).squeeze().cpu().numpy()
    return ai_prob, tag_probs


def write_music_header(csv_path):
    with open(csv_path, mode="w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "filename",
                "is_ai_generated", 
                "ai_confidence",
                "content_type",
                "genre",
                "mood", 
                "instruments",
                "tempo_bpm",
                "energy",
                "danceability"
            ]
        )

def get_basic_music_info(file_path: str) -> Dict[str, str]:
    """Fallback: basic info based on filename patterns"""
    filename = Path(file_path).name.lower()
    
    if any(word in filename for word in ['electronic', 'edm', 'house', 'techno']):
        genre = "Electronic"
        tempo = "128"
        energy = "High"
        mood = "Energetic"
    elif any(word in filename for word in ['rock', 'metal', 'punk']):
        genre = "Rock"
        tempo = "120"
        energy = "High"
        mood = "Energetic"
    elif any(word in filename for word in ['jazz', 'blues']):
        genre = "Jazz/Blues"
        tempo = "90"
        energy = "Medium"
        mood = "Smooth"
    elif any(word in filename for word in ['classical', 'symphony', 'concerto']):
        genre = "Classical"
        tempo = "80"
        energy = "Low"
        mood = "Calm"
    else:
        # Default for unknown music
        genre = "Popular Music"
        tempo = "100"
        energy = "Medium"
        mood = "Neutral"
    
    return {
        'genre': genre,
        'mood': mood,
        'instruments': 'Mixed',
        'tempo': tempo,
        'energy': energy,
        'danceability': '0.50'
    }


def analyze_music_simple(file_path: str) -> Dict[str, str]:
    """Simple music analysis that always works"""
    if LIBROSA_AVAILABLE:
        try:
            return analyze_with_librosa_safe(file_path)
        except Exception as e:
            print(f"Librosa analysis failed for {Path(file_path).name}: {str(e)[:50]}...")
            return get_basic_music_info(file_path)
    else:
        return get_basic_music_info(file_path)


def analyze_with_librosa_safe(file_path: str) -> Dict[str, str]:
    """Safe librosa analysis with error handling"""
    
    # Load audio (shorter duration to avoid memory issues)
    y, sr = librosa.load(file_path, duration=10, sr=22050)
    
    # Get tempo safely
    try:
        tempo, _ = librosa.beat.beat_track(y=y, sr=sr)
        tempo = float(tempo)  # Convert to Python float
    except:
        tempo = 100.0
    
    # Simple genre classification based on tempo
    if tempo > 140:
        genre = "Electronic/Dance"
        energy = "High"
        mood = "Energetic"
        danceability = 0.8
    elif tempo > 120:
        genre = "Pop/Rock"
        energy = "Medium"
        mood = "Energetic"
        danceability = 0.6
    elif tempo > 80:
        genre = "Folk/Alternative"
        energy = "Medium"
        mood = "Calm"
        danceability = 0.4
    else:
        genre = "Ballad/Classical"
        energy = "Low"
        mood = "Calm"
        danceability = 0.2
    
    return {
        'genre': genre,
        'mood': mood,
        'instruments': 'Mixed',
        'tempo': str(int(tempo)),  # Convert to string safely
        'energy': energy,
        'danceability': f"{danceability:.2f}"
    }

def append_music_row(
    csv_path: str,
    file_path: str,
    ai_label: Literal["Yes", "No"],
    ai_prob: Tensor,
    music_info: Dict[str, str],
) -> None:
    with open(csv_path, mode="a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        
        # Determine content type
        content_type = "AI-Generated Music" if ai_label == "Yes" else "Real Music"
        
        row = [
            file_path,
            ai_label, 
            f"{ai_prob:.3f}",
            content_type,
            music_info['genre'],
            music_info['mood'],
            music_info['instruments'], 
            music_info['tempo'],
            music_info['energy'],
            music_info['danceability']
        ]
        writer.writerow(row)
        
def get_audio_files(folder_path: str) -> list[Path]:
    """Get all audio files with case-insensitive extension matching"""
    folder = Path(folder_path)
    audio_files = []
    
    # Supported audio extensions (case-insensitive)
    supported_extensions = {'.wav', '.mp3', '.flac', '.m4a'}
    
    # Find all files and check extensions case-insensitively
    for file_path in folder.rglob("*"):
        if file_path.suffix.lower() in supported_extensions:
            audio_files.append(file_path)
    
    return audio_files
        
def write_final_accuracy_row(csv_path: str) -> str:
    """Write summary statistics to Excel file"""
    base_name = csv_path.replace('.csv', '')
    xlsx_path = f"{base_name}.xlsx"
    
    try:
        df = pd.read_csv(csv_path)
        
        # Calculate summary statistics
        total_files = len(df)
        ai_generated_count = len(df[df['is_ai_generated'] == 'Yes'])
        real_count = len(df[df['is_ai_generated'] == 'No'])
        avg_ai_confidence = df['ai_confidence'].mean()
        
        # Music-specific stats
        if 'genre' in df.columns:
            top_genres = df[df['is_ai_generated'] == 'No']['genre'].value_counts().head(3)
            top_genre = top_genres.index[0] if len(top_genres) > 0 else "N/A"
        else:
            top_genre = "N/A"
        
        # Write to Excel
        df.to_excel(xlsx_path, index=False)
        
        try:
            wb = load_workbook(xlsx_path)
            ws = wb.active
            
            # Add summary rows
            ws.append([])  # Blank row
            ws.append(['SUMMARY'])
            ws.append(['Total Files', total_files])
            ws.append(['Predicted AI Generated', ai_generated_count])
            ws.append(['Predicted Real', real_count])
            ws.append(['Average AI Confidence', f"{avg_ai_confidence:.3f}"])
            ws.append(['Top Genre (Real Music)', top_genre])
            
            # Highlight summary section
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            for row in range(ws.max_row - 5, ws.max_row + 1):
                for col in range(1, 3):  # First two columns
                    ws.cell(row=row, column=col).fill = yellow_fill
            
            wb.save(xlsx_path)
            
        except ImportError:
            print("Note: openpyxl not available for Excel formatting. Basic Excel file created.")
            
        return xlsx_path
        
    except Exception as e:
        print(f"Warning: Could not create Excel file: {e}")
        return csv_path
    
def predict_ai_only(model, input_tensor: Tensor) -> float:
    with torch.no_grad():
        if input_tensor.ndim == 2:
            input_tensor = input_tensor.squeeze(0)
        input_tensor = input_tensor.unsqueeze(0)
        binary_logit, _ = model(input_tensor)
        ai_prob = torch.sigmoid(binary_logit).item()
    return ai_prob
    
def predict_folder(folder_path: str, model_path: str, csv_path: str, threshold: float = 0.35):
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

    # Fixed: Use correct model class
    model = DualHeadCnn14Simple(pretrained=False)

    model.load_state_dict(torch.load(model_path, map_location=device))
    model.eval().to(device)

    write_music_header(csv_path)

    # Fixed: Use case-insensitive file finding
    audio_files_to_test = get_audio_files(folder_path)
    
    print(f"Found {len(audio_files_to_test)} audio files to process")
    
    # Show file types found
    extensions = {}
    for file_path in audio_files_to_test:
        ext = file_path.suffix.lower()
        extensions[ext] = extensions.get(ext, 0) + 1
    
    print("File types found:")
    for ext, count in extensions.items():
        print(f"  {ext}: {count} files")

    for file_path in tqdm(audio_files_to_test, desc="Running Prediction"):
        try:
            input_tensor = preprocess_audio(file_path=str(file_path)).to(device=device)
            ai_prob = predict_ai_only(model, input_tensor)
            ai_label = "Yes" if ai_prob > threshold else "No"
            
            music_info = analyze_music_simple(str(file_path))
            
            # If AI-generated, prefix the genre
            if ai_label == "Yes":
                music_info['genre'] = f"AI-{music_info['genre']}"
            
            # Write result with music tags
            append_music_row(csv_path, file_path.name, ai_label, ai_prob, music_info)
            
        except Exception as e:
            print(f"Error processing {file_path.name}: {e}")
            # Write error row
            error_info = {
                'genre': 'Processing Error',
                'mood': 'Unknown',
                'instruments': 'Unknown', 
                'tempo': 'Unknown',
                'energy': 'Unknown',
                'danceability': 'Unknown'
            }
            append_music_row(csv_path, file_path.name, "Unknown", 0.0, error_info)
            continue
    
    xlsx_path = write_final_accuracy_row(csv_path)
    print(f"Results saved to: {csv_path}")
    if xlsx_path != csv_path:
        print(f"Excel file created: {xlsx_path}")
        

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--folder",
        help="Folder containing .mp3/.wav/.flac/.m4a files to predict against",
        required=True,
    )
    parser.add_argument(
        "--model",
        help="Trained model in model/pretrained/saved_models/ or your own trained model",
        required=True,
    )
    parser.add_argument(
        "--output",
        help="Output Directory of where to put CSV file",
        required=True
    )
    args = parser.parse_args()
    csv_full_path = f"{args.output}/{DEFAULT_CSV_PATH}"
    
    print("AI AUDIO DETECTION FOR MUSIC")
    predict_folder(args.folder, args.model, csv_full_path)